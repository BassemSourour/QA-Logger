"""
Looks up roaster production data (pounds roasted and final temperature) for a
batch ticket from the two roaster log workbooks on the shared drive.

Why openpyxl instead of Excel COM automation:
- The roaster logs are large (6 MB and 18 MB) and live on a network share.
  Reading them cell-by-cell through COM took minutes and Excel answered with
  "Call was rejected by callee" while it was busy with the QA workbook.
- openpyxl in read_only mode streams the sheet XML and never touches the
  running Excel instance that owns the QA workbook.

Three things about these workbooks drive the implementation:
1. They contain chartsheets mixed in with the month sheets. A chartsheet has no
   rows, so it must be skipped rather than iterated.
2. Their reported dimensions are unreliable. The Roaster 2 July sheet claims
   max_row = 1048565 and max_column = 16384. Never trust max_row here; iterate
   with bounded guards instead.
3. Sheet naming is inconsistent across years ('2026 July' in one file,
   'JULY 2026 ' in the other), so month sheets are matched on tokens.

Layout, identical in both workbooks:
    row 6  = header
    row 7  = first data row
    col C  = BATCH TICKET
    col J  = FINAL TEMP.
    col M  = LBS

A batch ticket normally spans several roast rows. Pounds are summed across
those rows; the final temperature is taken from the last row that has one.
"""

import logging
import re
import threading
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROASTER_LOG_START_ROW = 7

BATCH_TICKET_COLUMN = 3
FINAL_TEMP_COLUMN = 10
LBS_COLUMN = 13

# Never read past the LBS column, and never trust the sheet's reported size.
MAX_COLUMN_TO_READ = LBS_COLUMN
MAX_SCAN_ROWS = 5000
MAX_CONSECUTIVE_BLANK_ROWS = 200

DEFAULT_CACHE_SECONDS = 120

# A lookup miss triggers one forced refresh, because a miss is exactly when the
# roast may have just been written to the log. This stops that refresh from
# running again on every miss in quick succession.
MIN_REFRESH_INTERVAL_SECONDS = 15

MONTH_TOKENS = {
    1: "JAN",
    2: "FEB",
    3: "MAR",
    4: "APR",
    5: "MAY",
    6: "JUN",
    7: "JUL",
    8: "AUG",
    9: "SEP",
    10: "OCT",
    11: "NOV",
    12: "DEC",
}


def normalize_batch_ticket_key(
    value,
) -> str:
    """
    Reduces every form a batch ticket appears in to one comparable key.

        'BT000096560' -> '96560'
        '000096560'   -> '96560'
        '96560.0'     -> '96560'
        96560.0       -> '96560'
        ' 96560 '     -> '96560'
    """
    if value is None:
        return ""

    text = str(
        value
    ).strip().upper().replace(
        " ",
        "",
    )

    if not text:
        return ""

    if text.endswith(".0"):
        text = text[:-2]

    if text.startswith("BT"):
        text = text[2:]

    if text.isdigit():
        return str(
            int(
                text
            )
        )

    return text


def parse_number(
    value,
):
    if value is None:
        return None

    if isinstance(
        value,
        bool,
    ):
        return None

    if isinstance(
        value,
        (
            int,
            float,
        ),
    ):
        return float(
            value
        )

    text = str(
        value
    ).strip().replace(
        ",",
        "",
    )

    if not text:
        return None

    try:
        return float(
            text
        )

    except ValueError:
        return None


def clean_number_for_excel(
    value,
):
    if value is None:
        return None

    number = float(
        value
    )

    if number.is_integer():
        return int(
            number
        )

    return round(
        number,
        2,
    )


def normalize_sheet_name_for_matching(
    sheet_name: str,
) -> str:
    return re.sub(
        r"[^A-Z0-9]",
        "",
        str(
            sheet_name
        ).upper(),
    )


def get_target_year_months(
    target_date: datetime,
) -> list[tuple[int, int]]:
    """
    Returns the current month and the previous month.

    A sample is usually QA tested the same day it is roasted, but a test can
    slip past a month boundary, so the previous month is included.
    """
    current_year = target_date.year
    current_month = target_date.month

    if current_month == 1:
        previous_year = current_year - 1
        previous_month = 12

    else:
        previous_year = current_year
        previous_month = current_month - 1

    return [
        (
            current_year,
            current_month,
        ),
        (
            previous_year,
            previous_month,
        ),
    ]


def find_month_sheet_names(
    sheet_names: list[str],
    wanted_year_months: list[tuple[int, int]],
) -> list[str]:
    matching_sheet_names = []

    for year, month in wanted_year_months:
        year_text = str(
            year
        )

        month_token = MONTH_TOKENS[
            month
        ]

        for sheet_name in sheet_names:
            normalized_name = normalize_sheet_name_for_matching(
                sheet_name
            )

            if year_text not in normalized_name:
                continue

            if month_token not in normalized_name:
                continue

            if sheet_name not in matching_sheet_names:
                matching_sheet_names.append(
                    sheet_name
                )

    return matching_sheet_names


def is_data_worksheet(
    worksheet,
) -> bool:
    """
    Chartsheets have no rows. Iterating one raises AttributeError on max_row.
    """
    return callable(
        getattr(
            worksheet,
            "iter_rows",
            None,
        )
    )


def get_row_value(
    row_values: tuple,
    column_number: int,
):
    index = column_number - 1

    if index < 0 or index >= len(
        row_values
    ):
        return None

    return row_values[
        index
    ]


def collect_rows_from_worksheet(
    worksheet,
    sheet_name: str,
    excel_path: Path,
    roaster_number,
    records: dict,
) -> int:
    """
    Accumulates one worksheet into records, keyed by
    (batch key, roaster number, sheet name).

    Keeping the sheet name in the key stops a duplicated month sheet (these
    workbooks contain pairs such as 'MARCH 2025' and 'MARCH 2025 (2)') from
    double counting the pounds for a batch.
    """
    rows_scanned = 0
    consecutive_blank_rows = 0

    row_iterator = worksheet.iter_rows(
        min_row=ROASTER_LOG_START_ROW,
        max_col=MAX_COLUMN_TO_READ,
        values_only=True,
    )

    for row_number, row_values in enumerate(
        row_iterator,
        start=ROASTER_LOG_START_ROW,
    ):
        rows_scanned += 1

        if rows_scanned > MAX_SCAN_ROWS:
            logging.warning(
                "Stopped scanning %s in %s after %s rows.",
                sheet_name,
                excel_path,
                MAX_SCAN_ROWS,
            )

            break

        batch_key = normalize_batch_ticket_key(
            get_row_value(
                row_values,
                BATCH_TICKET_COLUMN,
            )
        )

        if not batch_key:
            consecutive_blank_rows += 1

            if consecutive_blank_rows >= MAX_CONSECUTIVE_BLANK_ROWS:
                break

            continue

        consecutive_blank_rows = 0

        record_key = (
            batch_key,
            roaster_number,
            sheet_name,
        )

        record = records.get(
            record_key
        )

        if record is None:
            record = {
                "batch_ticket": batch_key,
                "quantity_roasted": None,
                "end_temperature": None,
                "roaster_number": roaster_number,
                "source_file": str(
                    excel_path
                ),
                "source_sheet": sheet_name,
                "source_rows": [],
            }

            records[
                record_key
            ] = record

        record[
            "source_rows"
        ].append(
            row_number
        )

        pounds_roasted = parse_number(
            get_row_value(
                row_values,
                LBS_COLUMN,
            )
        )

        if pounds_roasted is not None:
            if record["quantity_roasted"] is None:
                record["quantity_roasted"] = pounds_roasted

            else:
                record["quantity_roasted"] += pounds_roasted

        final_temperature = parse_number(
            get_row_value(
                row_values,
                FINAL_TEMP_COLUMN,
            )
        )

        if final_temperature is not None:
            record["end_temperature"] = final_temperature

    return rows_scanned


def read_one_roaster_log(
    roaster_log: dict,
    wanted_year_months: list[tuple[int, int]],
    records: dict,
) -> None:
    roaster_number = roaster_log.get(
        "roaster_number"
    )

    excel_path_text = str(
        roaster_log.get(
            "excel_path",
            "",
        )
    ).strip()

    if not excel_path_text:
        logging.warning(
            "Roaster log entry has no excel_path. Skipping it."
        )

        return

    excel_path = Path(
        excel_path_text
    )

    if not excel_path.exists():
        logging.warning(
            "Roaster log file not found: %s",
            excel_path,
        )

        return

    workbook = None
    started_at = time.time()

    try:
        logging.info(
            "Reading roaster log file: %s",
            excel_path,
        )

        workbook = load_workbook(
            filename=str(
                excel_path
            ),
            read_only=True,
            data_only=True,
        )

        sheet_names = find_month_sheet_names(
            sheet_names=workbook.sheetnames,
            wanted_year_months=wanted_year_months,
        )

        if not sheet_names:
            logging.warning(
                "No matching month sheets found in roaster log %s for %s",
                excel_path,
                wanted_year_months,
            )

            return

        total_rows_scanned = 0

        for sheet_name in sheet_names:
            worksheet = workbook[
                sheet_name
            ]

            if not is_data_worksheet(
                worksheet
            ):
                logging.info(
                    "Skipping chartsheet '%s' in %s",
                    sheet_name,
                    excel_path,
                )

                continue

            total_rows_scanned += collect_rows_from_worksheet(
                worksheet=worksheet,
                sheet_name=sheet_name,
                excel_path=excel_path,
                roaster_number=roaster_number,
                records=records,
            )

        logging.info(
            "Read roaster %s log in %.1f seconds | sheets=%s | rows=%s",
            roaster_number,
            time.time() - started_at,
            sheet_names,
            total_rows_scanned,
        )

    except Exception:
        logging.exception(
            "Failed to read roaster log file: %s",
            excel_path,
        )

    finally:
        if workbook is not None:
            try:
                workbook.close()

            except Exception:
                logging.exception(
                    "Failed to close roaster log workbook: %s",
                    excel_path,
                )


def collapse_records(
    records: dict,
) -> dict:
    """
    Reduces (batch key, roaster number, sheet name) records down to one record
    per batch key, preferring whichever source has the most roast rows.
    """
    best_by_batch_and_roaster = {}

    for (
        batch_key,
        roaster_number,
        _sheet_name,
    ), record in records.items():
        key = (
            batch_key,
            roaster_number,
        )

        existing = best_by_batch_and_roaster.get(
            key
        )

        if existing is None or len(
            record["source_rows"]
        ) > len(
            existing["source_rows"]
        ):
            best_by_batch_and_roaster[
                key
            ] = record

    lookup = {}

    for (
        batch_key,
        roaster_number,
    ), record in best_by_batch_and_roaster.items():
        existing = lookup.get(
            batch_key
        )

        if existing is None:
            lookup[
                batch_key
            ] = record

            continue

        logging.warning(
            "Batch ticket %s appears in more than one roaster log "
            "(roaster %s and roaster %s).",
            batch_key,
            existing["roaster_number"],
            roaster_number,
        )

        if len(
            record["source_rows"]
        ) > len(
            existing["source_rows"]
        ):
            lookup[
                batch_key
            ] = record

    for record in lookup.values():
        record["quantity_roasted"] = clean_number_for_excel(
            record["quantity_roasted"]
        )

        record["end_temperature"] = clean_number_for_excel(
            record["end_temperature"]
        )

    return lookup


def build_roaster_log_lookup(
    config: dict,
    target_date: datetime | None = None,
) -> dict:
    roaster_logs = config.get(
        "roaster_logs"
    ) or []

    if not roaster_logs:
        logging.info(
            "No roaster_logs configured. Roaster log lookup is disabled."
        )

        return {}

    if target_date is None:
        target_date = datetime.now()

    wanted_year_months = get_target_year_months(
        target_date
    )

    records = {}

    for roaster_log in roaster_logs:
        read_one_roaster_log(
            roaster_log=roaster_log,
            wanted_year_months=wanted_year_months,
            records=records,
        )

    lookup = collapse_records(
        records
    )

    logging.info(
        "Built roaster log lookup with %s batch ticket records.",
        len(
            lookup
        ),
    )

    return lookup


def lookup_roaster_log_data(
    lookup: dict,
    batch_ticket,
) -> dict | None:
    batch_key = normalize_batch_ticket_key(
        batch_ticket
    )

    if not batch_key:
        return None

    return lookup.get(
        batch_key
    )


class RoasterLogCache:
    """
    Holds the roaster log lookup so a barcode scan does not pay for a full
    read of both workbooks. Safe to call from more than one thread.
    """

    def __init__(
        self,
        config: dict,
    ) -> None:
        self._config = config
        self._state_lock = threading.Lock()
        self._build_lock = threading.Lock()
        self._lookup = {}
        self._loaded_at = 0.0

        cache_seconds = parse_number(
            config.get(
                "roaster_log_cache_seconds",
                DEFAULT_CACHE_SECONDS,
            )
        )

        if cache_seconds is None or cache_seconds <= 0:
            cache_seconds = DEFAULT_CACHE_SECONDS

        self._cache_seconds = float(
            cache_seconds
        )

    def is_enabled(
        self,
    ) -> bool:
        return bool(
            self._config.get(
                "roaster_logs"
            )
        )

    def refresh(
        self,
    ) -> dict:
        with self._build_lock:
            lookup = build_roaster_log_lookup(
                self._config
            )

            with self._state_lock:
                self._lookup = lookup
                self._loaded_at = time.time()

            return lookup

    def get_age_seconds(
        self,
    ) -> float:
        with self._state_lock:
            if self._loaded_at == 0.0:
                return float(
                    "inf"
                )

            return time.time() - self._loaded_at

    def get_lookup(
        self,
    ) -> dict:
        with self._state_lock:
            is_fresh = (
                self._loaded_at > 0.0
                and (
                    time.time() - self._loaded_at
                ) < self._cache_seconds
            )

            if is_fresh:
                return self._lookup

        return self.refresh()

    def get_roaster_log_data(
        self,
        batch_ticket,
    ) -> dict | None:
        if not self.is_enabled():
            return None

        lookup = self.get_lookup()

        roaster_log_data = lookup_roaster_log_data(
            lookup=lookup,
            batch_ticket=batch_ticket,
        )

        if roaster_log_data is not None:
            return roaster_log_data

        # A miss may simply mean the roast was logged after the last read.
        if self.get_age_seconds() < MIN_REFRESH_INTERVAL_SECONDS:
            return None

        logging.info(
            "Batch ticket %s not in the cached roaster log lookup. Refreshing.",
            batch_ticket,
        )

        lookup = self.refresh()

        return lookup_roaster_log_data(
            lookup=lookup,
            batch_ticket=batch_ticket,
        )
